"""
McLeuker Agentic AI Platform - Excel/CSV Execution Agent

Generates Excel and CSV files from structured data.
"""

import os
from typing import Optional, List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.agents.base_agent import BaseExecutionAgent
from src.models.schemas import (
    StructuredOutput,
    StructuredTable,
    GeneratedFile,
    OutputFormat
)


class ExcelAgent(BaseExecutionAgent):
    """
    Excel/CSV Execution Agent
    
    Generates Excel (.xlsx) and CSV files from structured table data.
    """
    
    @property
    def supported_formats(self) -> List[OutputFormat]:
        return [OutputFormat.EXCEL, OutputFormat.CSV]
    
    async def execute(
        self,
        structured_output: StructuredOutput,
        filename_prefix: Optional[str] = None
    ) -> List[GeneratedFile]:
        """
        Generate Excel/CSV files from structured output.
        
        Args:
            structured_output: The structured data containing tables
            filename_prefix: Optional prefix for filenames
            
        Returns:
            List[GeneratedFile]: List of generated files
        """
        generated_files = []
        
        if not structured_output.tables:
            return generated_files
        
        # Generate Excel file with all tables as sheets
        excel_file = await self._generate_excel(
            structured_output.tables,
            filename_prefix
        )
        if excel_file:
            generated_files.append(excel_file)
        
        # Generate individual CSV files for each table
        for table in structured_output.tables:
            csv_file = await self._generate_csv(table, filename_prefix)
            if csv_file:
                generated_files.append(csv_file)
        
        return generated_files
    
    async def _generate_excel(
        self,
        tables: List[StructuredTable],
        filename_prefix: Optional[str] = None
    ) -> Optional[GeneratedFile]:
        """Generate an Excel file with multiple sheets."""
        try:
            wb = Workbook()
            
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for table in tables:
                # Create sheet with sanitized name
                sheet_name = self._sanitize_sheet_name(table.name)
                ws = wb.create_sheet(title=sheet_name)
                
                # Create DataFrame from table data
                df = self._table_to_dataframe(table)
                
                # Write data to sheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = thin_border
                        
                        if r_idx == 1:
                            # Header row styling
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        else:
                            cell.alignment = cell_alignment
                
                # Adjust column widths
                for col_idx, column in enumerate(df.columns, 1):
                    max_length = max(
                        len(str(column)),
                        df[column].astype(str).map(len).max() if len(df) > 0 else 0
                    )
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
                
                # Freeze header row
                ws.freeze_panes = "A2"
            
            # Save workbook
            filename = self._generate_filename(filename_prefix, "xlsx")
            filepath = self._get_file_path(filename)
            wb.save(filepath)
            
            return self._create_generated_file(filename, OutputFormat.EXCEL)
            
        except Exception as e:
            print(f"Error generating Excel file: {e}")
            return None
    
    async def _generate_csv(
        self,
        table: StructuredTable,
        filename_prefix: Optional[str] = None
    ) -> Optional[GeneratedFile]:
        """Generate a CSV file from a single table."""
        try:
            df = self._table_to_dataframe(table)
            
            # Generate filename with table name
            suffix = self._sanitize_sheet_name(table.name).replace(" ", "_").lower()
            filename = self._generate_filename(filename_prefix, "csv", suffix)
            filepath = self._get_file_path(filename)
            
            df.to_csv(filepath, index=False, encoding="utf-8")
            
            return self._create_generated_file(filename, OutputFormat.CSV)
            
        except Exception as e:
            print(f"Error generating CSV file: {e}")
            return None
    
    def _table_to_dataframe(self, table: StructuredTable) -> pd.DataFrame:
        """Convert a StructuredTable to a pandas DataFrame."""
        if not table.rows:
            # Empty table with just columns
            return pd.DataFrame(columns=table.columns)
        
        # Handle different row formats
        rows_data = []
        for row in table.rows:
            if isinstance(row, dict):
                rows_data.append(row)
            elif isinstance(row, (list, tuple)):
                # Convert list to dict using columns
                row_dict = {}
                for i, col in enumerate(table.columns):
                    row_dict[col] = row[i] if i < len(row) else ""
                rows_data.append(row_dict)
        
        df = pd.DataFrame(rows_data)
        
        # Ensure all columns are present
        for col in table.columns:
            if col not in df.columns:
                df[col] = ""
        
        # Reorder columns to match table definition
        df = df[[col for col in table.columns if col in df.columns]]
        
        return df
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize a string for use as an Excel sheet name."""
        # Excel sheet names have restrictions
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, "_")
        
        # Limit to 31 characters (Excel limit)
        return sanitized[:31]
    
    async def generate_from_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str = "Data",
        filename_prefix: Optional[str] = None
    ) -> Optional[GeneratedFile]:
        """
        Generate an Excel file directly from a pandas DataFrame.
        
        Args:
            df: The DataFrame to convert
            sheet_name: Name for the Excel sheet
            filename_prefix: Optional prefix for filename
            
        Returns:
            GeneratedFile or None if failed
        """
        table = StructuredTable(
            name=sheet_name,
            description="",
            columns=list(df.columns),
            rows=df.to_dict('records')
        )
        
        return await self._generate_excel([table], filename_prefix)
