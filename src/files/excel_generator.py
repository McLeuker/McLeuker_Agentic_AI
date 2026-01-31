"""
McLeuker Agentic AI Platform - Professional Excel Generator

Creates professional, well-formatted Excel files with proper styling,
data validation, and structure like Manus AI.
"""

import os
import io
from typing import Optional, List, Dict, Any, Union
from datetime import datetime
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment,
    NamedStyle, Color
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from src.config.settings import get_settings


@dataclass
class ExcelColumn:
    """Definition of an Excel column."""
    name: str
    width: int = 15
    data_type: str = "string"  # string, number, currency, date, url, rating
    format_string: Optional[str] = None


@dataclass
class ExcelSheet:
    """Definition of an Excel sheet."""
    name: str
    data: List[Dict[str, Any]]
    columns: Optional[List[ExcelColumn]] = None
    title: Optional[str] = None
    description: Optional[str] = None
    include_chart: bool = False
    chart_type: str = "bar"  # bar, pie, line


@dataclass
class GeneratedExcelFile:
    """Result of Excel generation."""
    filename: str
    filepath: str
    size_bytes: int
    sheets: List[str]
    row_count: int
    created_at: datetime = field(default_factory=datetime.utcnow)
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "filepath": self.filepath,
            "size_bytes": self.size_bytes,
            "sheets": self.sheets,
            "row_count": self.row_count,
            "created_at": self.created_at.isoformat()
        }


class ProfessionalExcelGenerator:
    """
    Professional Excel file generator.
    
    Creates well-formatted, visually appealing Excel files with:
    - Professional styling and colors
    - Auto-sized columns
    - Data validation
    - Charts and visualizations
    - Proper data formatting
    """
    
    # Professional color scheme
    COLORS = {
        "primary": "1F4E79",      # Dark blue
        "secondary": "2E75B6",    # Medium blue
        "accent": "5B9BD5",       # Light blue
        "header_bg": "1F4E79",    # Dark blue
        "header_text": "FFFFFF",  # White
        "alt_row": "D6E3F8",      # Light blue tint
        "success": "70AD47",      # Green
        "warning": "FFC000",      # Yellow
        "error": "FF0000",        # Red
    }
    
    def __init__(self, output_dir: Optional[str] = None):
        self.settings = get_settings()
        self.output_dir = output_dir or self.settings.OUTPUT_DIR
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Define styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Set up reusable styles."""
        # Header style
        self.header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=self.COLORS["header_text"]
        )
        self.header_fill = PatternFill(
            start_color=self.COLORS["header_bg"],
            end_color=self.COLORS["header_bg"],
            fill_type="solid"
        )
        self.header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        
        # Data cell style
        self.data_font = Font(name='Calibri', size=10)
        self.data_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )
        
        # Border style
        self.thin_border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )
        
        # Alternate row fill
        self.alt_row_fill = PatternFill(
            start_color=self.COLORS["alt_row"],
            end_color=self.COLORS["alt_row"],
            fill_type="solid"
        )
    
    async def generate(
        self,
        sheets: List[ExcelSheet],
        filename: Optional[str] = None,
        title: Optional[str] = None
    ) -> GeneratedExcelFile:
        """
        Generate a professional Excel file.
        
        Args:
            sheets: List of sheet definitions
            filename: Optional filename (auto-generated if not provided)
            title: Optional workbook title
            
        Returns:
            GeneratedExcelFile with file details
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        total_rows = 0
        sheet_names = []
        
        for sheet_def in sheets:
            ws = wb.create_sheet(title=self._sanitize_sheet_name(sheet_def.name))
            sheet_names.append(ws.title)
            
            # Add title if provided
            start_row = 1
            if sheet_def.title:
                ws.merge_cells(f'A1:{get_column_letter(len(sheet_def.columns or []) or 5)}1')
                title_cell = ws['A1']
                title_cell.value = sheet_def.title
                title_cell.font = Font(name='Calibri', size=14, bold=True, color=self.COLORS["primary"])
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 30
                start_row = 2
            
            # Add description if provided
            if sheet_def.description:
                ws.merge_cells(f'A{start_row}:{get_column_letter(len(sheet_def.columns or []) or 5)}{start_row}')
                desc_cell = ws[f'A{start_row}']
                desc_cell.value = sheet_def.description
                desc_cell.font = Font(name='Calibri', size=10, italic=True)
                desc_cell.alignment = Alignment(horizontal="left", vertical="center")
                start_row += 1
            
            # Add empty row after title/description
            if sheet_def.title or sheet_def.description:
                start_row += 1
            
            # Write data
            if sheet_def.data:
                df = pd.DataFrame(sheet_def.data)
                rows_written = self._write_dataframe(ws, df, start_row, sheet_def.columns)
                total_rows += rows_written
                
                # Add chart if requested
                if sheet_def.include_chart and len(df) > 0:
                    self._add_chart(ws, df, start_row, sheet_def.chart_type)
        
        # Generate filename
        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"data_{timestamp}.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Save workbook
        wb.save(filepath)
        
        # Get file size
        size_bytes = os.path.getsize(filepath)
        
        return GeneratedExcelFile(
            filename=filename,
            filepath=filepath,
            size_bytes=size_bytes,
            sheets=sheet_names,
            row_count=total_rows
        )
    
    def _write_dataframe(
        self,
        ws,
        df: pd.DataFrame,
        start_row: int,
        columns: Optional[List[ExcelColumn]] = None
    ) -> int:
        """Write a DataFrame to a worksheet with professional formatting."""
        if df.empty:
            return 0
        
        # Determine column definitions
        col_defs = {}
        if columns:
            for col in columns:
                col_defs[col.name] = col
        
        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(column))
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Set header row height
        ws.row_dimensions[start_row].height = 25
        
        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            is_alt_row = (row_idx - start_row) % 2 == 0
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Get column definition
                col_name = df.columns[col_idx - 1]
                col_def = col_defs.get(col_name)
                
                # Format value based on type
                cell.value = self._format_cell_value(value, col_def)
                
                # Apply styling
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.thin_border
                
                if is_alt_row:
                    cell.fill = self.alt_row_fill
                
                # Apply number format if specified
                if col_def and col_def.format_string:
                    cell.number_format = col_def.format_string
                elif col_def:
                    if col_def.data_type == "currency":
                        cell.number_format = '$#,##0.00'
                    elif col_def.data_type == "number":
                        cell.number_format = '#,##0'
                    elif col_def.data_type == "date":
                        cell.number_format = 'YYYY-MM-DD'
        
        # Auto-size columns
        self._auto_size_columns(ws, df, columns)
        
        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        
        # Add filters
        if len(df) > 0:
            last_col = get_column_letter(len(df.columns))
            last_row = start_row + len(df)
            ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"
        
        return len(df)
    
    def _format_cell_value(self, value: Any, col_def: Optional[ExcelColumn] = None) -> Any:
        """Format a cell value based on its type."""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        
        if col_def:
            if col_def.data_type == "rating":
                # Convert rating to stars
                try:
                    rating = float(value)
                    full_stars = int(rating)
                    half_star = rating - full_stars >= 0.5
                    return "★" * full_stars + ("½" if half_star else "") + f" ({rating})"
                except (ValueError, TypeError):
                    return str(value)
            
            elif col_def.data_type == "url":
                return str(value)
        
        return value
    
    def _auto_size_columns(
        self,
        ws,
        df: pd.DataFrame,
        columns: Optional[List[ExcelColumn]] = None
    ):
        """Auto-size columns based on content."""
        col_defs = {}
        if columns:
            for col in columns:
                col_defs[col.name] = col
        
        for col_idx, column in enumerate(df.columns, 1):
            col_def = col_defs.get(column)
            
            # Use defined width or calculate
            if col_def and col_def.width:
                width = col_def.width
            else:
                # Calculate width based on content
                max_length = len(str(column))
                
                for value in df[column]:
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
                
                width = min(max_length + 2, 50)
            
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def _add_chart(
        self,
        ws,
        df: pd.DataFrame,
        data_start_row: int,
        chart_type: str = "bar"
    ):
        """Add a chart to the worksheet."""
        if len(df) < 2 or len(df.columns) < 2:
            return
        
        # Find numeric columns for chart
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if not numeric_cols:
            return
        
        # Create chart
        if chart_type == "pie":
            chart = PieChart()
        elif chart_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()
        
        chart.title = "Data Visualization"
        chart.style = 10
        
        # Set data reference
        data_end_row = data_start_row + len(df)
        
        # Use first numeric column for values
        value_col = df.columns.get_loc(numeric_cols[0]) + 1
        
        data = Reference(
            ws,
            min_col=value_col,
            min_row=data_start_row,
            max_row=data_end_row,
            max_col=value_col
        )
        
        # Use first column for categories
        cats = Reference(
            ws,
            min_col=1,
            min_row=data_start_row + 1,
            max_row=data_end_row
        )
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Position chart
        chart_col = len(df.columns) + 2
        ws.add_chart(chart, f"{get_column_letter(chart_col)}{data_start_row}")
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize a string for use as an Excel sheet name."""
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, "_")
        return sanitized[:31]
    
    async def generate_from_data(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        filename: Optional[str] = None,
        title: Optional[str] = None,
        description: Optional[str] = None
    ) -> GeneratedExcelFile:
        """
        Generate Excel from a simple data list.
        
        Args:
            data: List of dictionaries
            sheet_name: Name for the sheet
            filename: Optional filename
            title: Optional title for the sheet
            description: Optional description
            
        Returns:
            GeneratedExcelFile
        """
        if not data:
            # Create empty file with headers
            data = [{}]
        
        # Auto-detect columns
        all_keys = set()
        for row in data:
            all_keys.update(row.keys())
        
        columns = [ExcelColumn(name=key, width=20) for key in sorted(all_keys)]
        
        sheet = ExcelSheet(
            name=sheet_name,
            data=data,
            columns=columns,
            title=title,
            description=description
        )
        
        return await self.generate([sheet], filename)
    
    async def generate_comparison_table(
        self,
        items: List[Dict[str, Any]],
        comparison_fields: List[str],
        filename: Optional[str] = None,
        title: str = "Comparison"
    ) -> GeneratedExcelFile:
        """
        Generate a comparison table Excel file.
        
        Args:
            items: List of items to compare
            comparison_fields: Fields to include in comparison
            filename: Optional filename
            title: Table title
            
        Returns:
            GeneratedExcelFile
        """
        # Restructure data for comparison format
        comparison_data = []
        
        for item in items:
            row = {}
            for field in comparison_fields:
                row[field] = item.get(field, "N/A")
            comparison_data.append(row)
        
        return await self.generate_from_data(
            comparison_data,
            sheet_name="Comparison",
            filename=filename,
            title=title
        )


# Global generator instance
_excel_generator: Optional[ProfessionalExcelGenerator] = None


def get_excel_generator(output_dir: Optional[str] = None) -> ProfessionalExcelGenerator:
    """Get or create the global Excel generator."""
    global _excel_generator
    if _excel_generator is None:
        _excel_generator = ProfessionalExcelGenerator(output_dir)
    return _excel_generator
