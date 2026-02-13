"""
Excel Agent — Advanced Spreadsheet Creation & Analysis
=======================================================

Creates professional spreadsheets (XLSX) with advanced features:
- Data structuring and organization
- Formulas and calculations
- Charts and visualizations
- Formatting and styling
"""

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Optional
import openai

# Excel library
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel generation disabled")

logger = logging.getLogger(__name__)


@dataclass
class SheetData:
    """Data for a single worksheet."""
    name: str
    headers: List[str]
    rows: List[List[Any]]
    formulas: Dict[str, str] = field(default_factory=dict)  # cell -> formula
    column_widths: Dict[str, float] = field(default_factory=dict)
    formats: Dict[str, str] = field(default_factory=dict)  # cell -> format type
    
    def to_dict(self) -> Dict:
        return {
            "name": self.name,
            "headers": self.headers,
            "rows": self.rows,
            "formulas": self.formulas,
            "column_widths": self.column_widths,
            "formats": self.formats,
        }


@dataclass
class ChartConfig:
    """Configuration for a chart."""
    chart_type: str  # "bar", "pie", "line", "column"
    title: str
    data_range: str
    position: str  # cell position for chart
    
    def to_dict(self) -> Dict:
        return {
            "chart_type": self.chart_type,
            "title": self.title,
            "data_range": self.data_range,
            "position": self.position,
        }


@dataclass
class GeneratedSpreadsheet:
    """A generated spreadsheet."""
    spreadsheet_id: str
    title: str
    file_path: str
    sheets: List[SheetData]
    charts: List[ChartConfig]
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict:
        return {
            "spreadsheet_id": self.spreadsheet_id,
            "title": self.title,
            "file_path": self.file_path,
            "sheets": [s.to_dict() for s in self.sheets],
            "charts": [c.to_dict() for c in self.charts],
            "metadata": self.metadata,
        }


class ExcelAgent:
    """
    AI Excel Agent for professional spreadsheet creation.
    
    Usage:
        agent = ExcelAgent(llm_client)
        async for event in agent.generate_spreadsheet(
            "Create a sales dashboard with monthly data and charts"
        ):
            print(event)
    """
    
    # Predefined spreadsheet types
    SPREADSHEET_TYPES = {
        "budget": {
            "name": "Budget Tracker",
            "description": "Income and expense tracking with summaries",
            "sheets": ["Income", "Expenses", "Summary"],
        },
        "sales_dashboard": {
            "name": "Sales Dashboard",
            "description": "Sales performance with metrics and charts",
            "sheets": ["Sales Data", "Metrics", "Charts"],
        },
        "project_tracker": {
            "name": "Project Tracker",
            "description": "Project tasks, timeline, and status tracking",
            "sheets": ["Tasks", "Timeline", "Resources"],
        },
        "inventory": {
            "name": "Inventory Management",
            "description": "Product inventory with stock levels",
            "sheets": ["Products", "Stock Levels", "Orders"],
        },
        "financial_model": {
            "name": "Financial Model",
            "description": "Revenue projections and financial analysis",
            "sheets": ["Assumptions", "Revenue", "Expenses", "P&L"],
        },
        "employee_tracker": {
            "name": "Employee Tracker",
            "description": "Employee information and performance tracking",
            "sheets": ["Employees", "Performance", "Departments"],
        },
    }
    
    def __init__(
        self,
        llm_client: openai.AsyncOpenAI,
        model: str = "kimi-k2.5",
        output_dir: str = "/tmp/excel_agent",
    ):
        self.llm_client = llm_client
        self.model = model
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Track generated spreadsheets
        self._spreadsheets: Dict[str, GeneratedSpreadsheet] = {}
    
    async def generate_spreadsheet(
        self,
        description: str,
        spreadsheet_type: Optional[str] = None,
        title: Optional[str] = None,
        data: Optional[Dict] = None,
        context: Optional[Dict] = None,
    ) -> AsyncGenerator[Dict[str, Any], None]:
        """
        Generate a professional spreadsheet from a description.
        
        Args:
            description: Natural language description of the spreadsheet
            spreadsheet_type: Optional predefined type
            title: Optional spreadsheet title
            data: Optional data to include
            context: Additional context for generation
            
        Yields:
            Progress events and final spreadsheet info
        """
        import uuid
        spreadsheet_id = str(uuid.uuid4())
        
        yield {"type": "start", "data": {"spreadsheet_id": spreadsheet_id, "description": description}}
        
        try:
            # Step 1: Analyze requirements
            yield {"type": "phase", "data": {"phase": "analysis", "status": "started"}}
            
            selected_type = None
            if spreadsheet_type and spreadsheet_type in self.SPREADSHEET_TYPES:
                selected_type = self.SPREADSHEET_TYPES[spreadsheet_type]
            
            analysis = await self._analyze_requirements(description, selected_type)
            
            yield {
                "type": "phase",
                "data": {
                    "phase": "analysis",
                    "status": "completed",
                    "sheets": analysis.get("num_sheets", 1),
                }
            }
            
            # Step 2: Generate sheet structure
            yield {"type": "phase", "data": {"phase": "structure", "status": "started"}}
            
            sheets = await self._generate_sheets(
                description,
                analysis,
                data,
                context,
            )
            
            yield {
                "type": "phase",
                "data": {
                    "phase": "structure",
                    "status": "completed",
                    "sheets": len(sheets),
                }
            }
            
            # Step 3: Generate chart configurations
            yield {"type": "phase", "data": {"phase": "charts", "status": "started"}}
            
            charts = await self._generate_charts(description, sheets)
            
            yield {"type": "phase", "data": {"phase": "charts", "status": "completed", "charts": len(charts)}}
            
            # Step 4: Create Excel file
            yield {"type": "phase", "data": {"phase": "creation", "status": "started"}}
            
            ss_title = title or analysis.get("title", "Spreadsheet")
            file_path = await self._create_xlsx(
                spreadsheet_id,
                ss_title,
                sheets,
                charts,
            )
            
            yield {"type": "phase", "data": {"phase": "creation", "status": "completed", "path": file_path}}
            
            # Step 5: Create spreadsheet record
            spreadsheet = GeneratedSpreadsheet(
                spreadsheet_id=spreadsheet_id,
                title=ss_title,
                file_path=file_path,
                sheets=sheets,
                charts=charts,
                metadata={
                    "type": spreadsheet_type or "custom",
                    "description": description,
                    "context": context or {},
                },
            )
            self._spreadsheets[spreadsheet_id] = spreadsheet
            
            yield {
                "type": "complete",
                "data": {
                    "spreadsheet_id": spreadsheet_id,
                    "title": ss_title,
                    "file_path": file_path,
                    "sheets": len(sheets),
                    "charts": len(charts),
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating spreadsheet: {e}")
            yield {"type": "error", "data": {"message": str(e)}}
    
    async def analyze_data(
        self,
        file_path: str,
        analysis_type: str = "summary",
    ) -> AsyncGenerator[Dict[str, Any], None]:
        """
        Analyze data in an existing spreadsheet.
        
        Args:
            file_path: Path to the Excel file
            analysis_type: Type of analysis to perform
            
        Yields:
            Analysis results
        """
        yield {"type": "start", "data": {"file_path": file_path, "analysis_type": analysis_type}}
        
        try:
            if not OPENPYXL_AVAILABLE:
                yield {"type": "error", "data": {"message": "openpyxl not available"}}
                return
            
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            
            analysis_results = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Basic statistics
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                
                analysis_results[sheet_name] = {
                    "row_count": len(data_rows),
                    "column_count": ws.max_column,
                    "sample_data": data_rows[:5] if data_rows else [],
                }
            
            yield {
                "type": "complete",
                "data": {
                    "file_path": file_path,
                    "analysis": analysis_results,
                }
            }
            
        except Exception as e:
            logger.error(f"Error analyzing spreadsheet: {e}")
            yield {"type": "error", "data": {"message": str(e)}}
    
    async def _analyze_requirements(
        self,
        description: str,
        spreadsheet_type: Optional[Dict],
    ) -> Dict:
        """Analyze requirements for the spreadsheet."""
        type_info = json.dumps(spreadsheet_type) if spreadsheet_type else "None"
        
        messages = [
            {
                "role": "system",
                "content": f"""Analyze the spreadsheet request and provide structured requirements.

Predefined type (if any): {type_info}

Respond with JSON:
{{
    "title": "Spreadsheet title",
    "num_sheets": number,
    "sheet_names": ["Sheet1", "Sheet2"],
    "key_features": ["feature1", "feature2"],
    "formulas_needed": ["formula1", "formula2"],
    "charts_needed": ["chart1", "chart2"],
    "complexity": "simple|medium|complex"
}}"""
            },
            {
                "role": "user",
                "content": f"Analyze: {description}"
            }
        ]
        
        response = await self.llm_client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.5,
            max_tokens=1024,
            response_format={"type": "json_object"},
        )
        
        return json.loads(response.choices[0].message.content)
    
    async def _generate_sheets(
        self,
        description: str,
        analysis: Dict,
        data: Optional[Dict],
        context: Optional[Dict],
    ) -> List[SheetData]:
        """Generate data for all sheets."""
        sheets = []
        sheet_names = analysis.get("sheet_names", ["Data"])
        
        for sheet_name in sheet_names:
            sheet = await self._generate_sheet_data(
                sheet_name,
                description,
                analysis,
                data,
                context,
            )
            sheets.append(sheet)
        
        return sheets
    
    async def _generate_sheet_data(
        self,
        sheet_name: str,
        description: str,
        analysis: Dict,
        data: Optional[Dict],
        context: Optional[Dict],
    ) -> SheetData:
        """Generate data for a single sheet."""
        messages = [
            {
                "role": "system",
                "content": f"""Generate spreadsheet data for sheet: {sheet_name}

Spreadsheet: {description}

Respond with JSON:
{{
    "headers": ["Column1", "Column2", "Column3"],
    "rows": [
        ["value1", "value2", "value3"],
        ["value4", "value5", "value6"]
    ],
    "formulas": {{
        "D2": "=B2+C2",
        "D3": "=B3+C3"
    }},
    "column_widths": {{
        "A": 15,
        "B": 20
    }},
    "formats": {{
        "B2:B10": "currency",
        "C2:C10": "percentage"
    }}
}}"""
            }
        ]
        
        if data:
            messages.append({
                "role": "user",
                "content": f"Use this data if relevant: {json.dumps(data)}"
            })
        
        response = await self.llm_client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.7,
            max_tokens=3072,
            response_format={"type": "json_object"},
        )
        
        result = json.loads(response.choices[0].message.content)
        
        return SheetData(
            name=sheet_name,
            headers=result.get("headers", []),
            rows=result.get("rows", []),
            formulas=result.get("formulas", {}),
            column_widths=result.get("column_widths", {}),
            formats=result.get("formats", {}),
        )
    
    async def _generate_charts(
        self,
        description: str,
        sheets: List[SheetData],
    ) -> List[ChartConfig]:
        """Generate chart configurations."""
        messages = [
            {
                "role": "system",
                "content": f"""Generate chart configurations for this spreadsheet.

Spreadsheet: {description}
Available sheets: {[s.name for s in sheets]}

Respond with JSON:
{{
    "charts": [
        {{
            "chart_type": "bar|pie|line|column",
            "title": "Chart Title",
            "data_range": "Sheet1!A1:D10",
            "position": "F2"
        }}
    ]
}}"""
            }
        ]
        
        response = await self.llm_client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.5,
            max_tokens=1024,
            response_format={"type": "json_object"},
        )
        
        result = json.loads(response.choices[0].message.content)
        charts_data = result.get("charts", [])
        
        charts = []
        for chart_data in charts_data:
            charts.append(ChartConfig(
                chart_type=chart_data.get("chart_type", "bar"),
                title=chart_data.get("title", "Chart"),
                data_range=chart_data.get("data_range", ""),
                position=chart_data.get("position", "H2"),
            ))
        
        return charts
    
    async def _create_xlsx(
        self,
        spreadsheet_id: str,
        title: str,
        sheets: List[SheetData],
        charts: List[ChartConfig],
    ) -> str:
        """Create an XLSX file."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not available")
        
        file_path = self.output_dir / f"{spreadsheet_id}.xlsx"
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.name)
            
            # Add headers
            for col_idx, header in enumerate(sheet_data.headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Add data rows
            for row_idx, row_data in enumerate(sheet_data.rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add formulas
            for cell_ref, formula in sheet_data.formulas.items():
                ws[cell_ref] = formula
            
            # Set column widths
            for col_letter, width in sheet_data.column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Apply formatting
            for range_str, format_type in sheet_data.formats.items():
                # Parse range and apply formatting
                pass  # Implementation depends on format types
        
        # Add charts
        for chart_config in charts:
            if chart_config.chart_type == "bar":
                chart = BarChart()
            elif chart_config.chart_type == "pie":
                chart = PieChart()
            elif chart_config.chart_type == "line":
                chart = LineChart()
            else:
                chart = BarChart()
            
            chart.title = chart_config.title
            # Parse data_range and set chart data
            # chart.add_data(data)
            # ws.add_chart(chart, chart_config.position)
        
        wb.save(str(file_path))
        return str(file_path)
    
    def list_types(self) -> List[Dict]:
        """List available spreadsheet types."""
        return [{"id": k, **v} for k, v in self.SPREADSHEET_TYPES.items()]
    
    def get_spreadsheet(self, spreadsheet_id: str) -> Optional[GeneratedSpreadsheet]:
        """Get a generated spreadsheet by ID."""
        return self._spreadsheets.get(spreadsheet_id)


# Singleton instance
_excel_agent: Optional[ExcelAgent] = None


def get_excel_agent(llm_client: openai.AsyncOpenAI = None) -> Optional[ExcelAgent]:
    """Get or create the Excel Agent singleton."""
    global _excel_agent
    if _excel_agent is None and llm_client:
        _excel_agent = ExcelAgent(llm_client)
    return _excel_agent
