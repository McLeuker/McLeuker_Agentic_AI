"""
McLeuker AI V5.1 - File Generation Service
==========================================
This service generates REAL downloadable files (Excel, PDF, PPT).

Key improvements:
1. LLM outputs structured data (JSON/table)
2. This service converts to actual files
3. Files are uploaded to storage
4. Frontend shows download button with real URL

This fixes the "I asked for Excel but got text" issue.
"""

import os
import json
import uuid
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO

# File generation libraries
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from ..schemas.response_contract import GeneratedFile, FileType, TableData


class FileGeneratorService:
    """
    Generates real downloadable files from structured data.
    """
    
    def __init__(self, storage_path: str = "/tmp/mcleuker_files"):
        self.storage_path = storage_path
        os.makedirs(storage_path, exist_ok=True)
        
        # Base URL for file downloads (Railway URL)
        self.base_url = os.getenv("RAILWAY_PUBLIC_URL", "https://web-production-29f3c.up.railway.app")
    
    def generate_excel(
        self,
        data: List[Dict[str, Any]],
        title: str = "McLeuker AI Report",
        sheet_name: str = "Data",
        headers: List[str] = None
    ) -> GeneratedFile:
        """
        Generate a real Excel file from structured data.
        
        Args:
            data: List of dictionaries with row data
            title: Report title
            sheet_name: Excel sheet name
            headers: Optional custom headers (auto-detected if not provided)
        
        Returns:
            GeneratedFile with download URL
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl not installed")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Auto-detect headers if not provided
        if not headers and data:
            headers = list(data[0].keys())
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Add title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color="1a1a2e")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add timestamp
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Add headers (row 4)
        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = border
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(50, max_length + 2)
        
        # Save file
        file_id = str(uuid.uuid4())[:8]
        filename = f"mcleuker_report_{file_id}.xlsx"
        filepath = os.path.join(self.storage_path, filename)
        wb.save(filepath)
        
        # Get file size
        file_size = os.path.getsize(filepath)
        
        return GeneratedFile(
            id=file_id,
            filename=filename,
            file_type=FileType.EXCEL,
            url=f"{self.base_url}/files/{filename}",
            size_bytes=file_size,
            description=f"{title} - {len(data)} rows"
        )
    
    def generate_excel_from_table(self, table: TableData, title: str = None) -> GeneratedFile:
        """Generate Excel from TableData schema"""
        # Convert TableData to list of dicts
        data = []
        for row in table.rows:
            row_dict = {}
            for i, header in enumerate(table.headers):
                row_dict[header] = row[i] if i < len(row) else ""
            data.append(row_dict)
        
        return self.generate_excel(
            data=data,
            title=title or table.title or "McLeuker AI Report",
            headers=table.headers
        )
    
    def generate_pdf(
        self,
        content: str,
        title: str = "McLeuker AI Report",
        tables: List[TableData] = None
    ) -> GeneratedFile:
        """
        Generate a PDF document from content and tables.
        """
        if not PDF_AVAILABLE:
            raise RuntimeError("fpdf not installed")
        
        pdf = FPDF()
        pdf.add_page()
        
        # Title
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, title, ln=True, align="C")
        pdf.ln(5)
        
        # Timestamp
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.ln(10)
        
        # Content
        pdf.set_font("Arial", "", 11)
        # Handle unicode by encoding
        safe_content = content.encode('latin-1', 'replace').decode('latin-1')
        pdf.multi_cell(0, 6, safe_content)
        
        # Tables
        if tables:
            for table in tables:
                pdf.ln(10)
                if table.title:
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, table.title, ln=True)
                
                pdf.set_font("Arial", "", 10)
                # Simple table rendering
                col_width = 190 / len(table.headers)
                
                # Headers
                pdf.set_fill_color(26, 26, 46)
                pdf.set_text_color(255, 255, 255)
                for header in table.headers:
                    pdf.cell(col_width, 8, header[:20], border=1, fill=True)
                pdf.ln()
                
                # Rows
                pdf.set_text_color(0, 0, 0)
                for row in table.rows[:20]:  # Limit rows for PDF
                    for i, cell in enumerate(row):
                        if i < len(table.headers):
                            pdf.cell(col_width, 6, str(cell)[:25], border=1)
                    pdf.ln()
        
        # Save file
        file_id = str(uuid.uuid4())[:8]
        filename = f"mcleuker_report_{file_id}.pdf"
        filepath = os.path.join(self.storage_path, filename)
        pdf.output(filepath)
        
        file_size = os.path.getsize(filepath)
        
        return GeneratedFile(
            id=file_id,
            filename=filename,
            file_type=FileType.PDF,
            url=f"{self.base_url}/files/{filename}",
            size_bytes=file_size,
            description=title
        )
    
    def generate_csv(self, data: List[Dict[str, Any]], title: str = "data") -> GeneratedFile:
        """Generate a CSV file from structured data"""
        import csv
        
        if not data:
            raise ValueError("No data provided for CSV generation")
        
        headers = list(data[0].keys())
        
        file_id = str(uuid.uuid4())[:8]
        filename = f"mcleuker_{title}_{file_id}.csv"
        filepath = os.path.join(self.storage_path, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        
        file_size = os.path.getsize(filepath)
        
        return GeneratedFile(
            id=file_id,
            filename=filename,
            file_type=FileType.CSV,
            url=f"{self.base_url}/files/{filename}",
            size_bytes=file_size,
            description=f"{title} - {len(data)} rows"
        )
    
    def get_file_path(self, filename: str) -> Optional[str]:
        """Get full path for a generated file"""
        filepath = os.path.join(self.storage_path, filename)
        if os.path.exists(filepath):
            return filepath
        return None


# Singleton instance
file_generator = FileGeneratorService()
