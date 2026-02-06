"""
V3.1 Analyst Layer
Code execution using E2B sandbox for data analysis and professional file generation.
"""

import asyncio
import aiohttp
import json
import base64
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
import os

from src.config.settings import settings


@dataclass
class AnalystResult:
    """Result from an analyst operation."""
    success: bool
    analysis: str
    files: List[Dict]
    error: Optional[str] = None


class E2BSandbox:
    """
    E2B Sandbox integration for secure Python code execution.
    Used for data analysis, Excel generation, and complex calculations.
    """
    
    def __init__(self):
        self.api_key = settings.E2B_API_KEY
        self.enabled = bool(self.api_key)
        self.base_url = "https://api.e2b.dev/v1"
    
    async def execute_code(self, code: str, timeout: int = None) -> Dict:
        """Execute Python code in a secure sandbox."""
        if not self.enabled:
            return {"success": False, "error": "E2B not configured"}
        
        timeout = timeout or settings.E2B_TIMEOUT
        
        try:
            # Create a new sandbox session
            async with aiohttp.ClientSession() as session:
                # Start sandbox
                create_resp = await session.post(
                    f"{self.base_url}/sandboxes",
                    headers={
                        "X-API-Key": self.api_key,
                        "Content-Type": "application/json"
                    },
                    json={"template": "base"},
                    timeout=aiohttp.ClientTimeout(total=30)
                )
                
                if create_resp.status != 200 and create_resp.status != 201:
                    error_text = await create_resp.text()
                    return {"success": False, "error": f"Failed to create sandbox: {error_text}"}
                
                sandbox_data = await create_resp.json()
                sandbox_id = sandbox_data.get("sandboxId") or sandbox_data.get("id")
                
                if not sandbox_id:
                    return {"success": False, "error": "No sandbox ID returned"}
                
                # Execute code
                exec_resp = await session.post(
                    f"{self.base_url}/sandboxes/{sandbox_id}/executions",
                    headers={
                        "X-API-Key": self.api_key,
                        "Content-Type": "application/json"
                    },
                    json={
                        "code": code,
                        "language": "python"
                    },
                    timeout=aiohttp.ClientTimeout(total=timeout)
                )
                
                if exec_resp.status != 200:
                    error_text = await exec_resp.text()
                    return {"success": False, "error": f"Execution failed: {error_text}"}
                
                result = await exec_resp.json()
                
                # Clean up sandbox
                await session.delete(
                    f"{self.base_url}/sandboxes/{sandbox_id}",
                    headers={"X-API-Key": self.api_key}
                )
                
                return {
                    "success": True,
                    "stdout": result.get("stdout", ""),
                    "stderr": result.get("stderr", ""),
                    "result": result.get("result"),
                    "files": result.get("files", [])
                }
                
        except asyncio.TimeoutError:
            return {"success": False, "error": "Execution timeout"}
        except Exception as e:
            return {"success": False, "error": str(e)}


class ExcelGenerator:
    """
    Professional Excel file generator using E2B sandbox.
    Creates stunning, data-rich spreadsheets.
    """
    
    def __init__(self, sandbox: E2BSandbox):
        self.sandbox = sandbox
    
    async def generate(self, data: Dict, title: str = "Report") -> Dict:
        """Generate a professional Excel file from data."""
        
        # Build the Python code for Excel generation
        code = f'''
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import json
import base64

# Data from the agent
data = {json.dumps(data)}
title = "{title}"

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Styling
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:E1')
ws['A1'] = title
ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws['A1'].alignment = Alignment(horizontal='center')

# Add data
if isinstance(data, dict):
    if 'rows' in data and 'columns' in data:
        # Structured table data
        df = pd.DataFrame(data['rows'], columns=data['columns'])
    elif 'items' in data:
        # List of items
        df = pd.DataFrame(data['items'])
    else:
        # Key-value pairs
        df = pd.DataFrame(list(data.items()), columns=['Key', 'Value'])
elif isinstance(data, list):
    df = pd.DataFrame(data)
else:
    df = pd.DataFrame({{'Data': [str(data)]}})

# Write headers
for col_idx, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=3, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Write data
for row_idx, row in enumerate(df.values, 4):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(horizontal='left')
        
        # Alternate row colors
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Save to bytes
import io
output = io.BytesIO()
wb.save(output)
output.seek(0)

# Return base64 encoded file
result = base64.b64encode(output.read()).decode('utf-8')
print(f"EXCEL_OUTPUT:{{result}}")
'''
        
        result = await self.sandbox.execute_code(code)
        
        if not result.get("success"):
            return {"success": False, "error": result.get("error")}
        
        # Extract the base64 file from stdout
        stdout = result.get("stdout", "")
        if "EXCEL_OUTPUT:" in stdout:
            file_data = stdout.split("EXCEL_OUTPUT:")[1].strip()
            return {
                "success": True,
                "file_data": file_data,
                "filename": f"{title.replace(' ', '_')}.xlsx",
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        
        return {"success": False, "error": "Failed to generate Excel file"}


class DataAnalyzer:
    """
    Data analysis engine using E2B sandbox.
    Performs calculations, statistics, and data transformations.
    """
    
    def __init__(self, sandbox: E2BSandbox):
        self.sandbox = sandbox
    
    async def analyze(self, data: Any, analysis_type: str = "summary") -> Dict:
        """Analyze data and return insights."""
        
        code = f'''
import pandas as pd
import json

data = {json.dumps(data)}

# Convert to DataFrame
if isinstance(data, dict):
    if 'rows' in data:
        df = pd.DataFrame(data['rows'], columns=data.get('columns', None))
    else:
        df = pd.DataFrame([data])
elif isinstance(data, list):
    df = pd.DataFrame(data)
else:
    df = pd.DataFrame({{'value': [data]}})

# Analysis
analysis = {{
    "shape": df.shape,
    "columns": list(df.columns),
    "dtypes": df.dtypes.astype(str).to_dict(),
    "summary": df.describe().to_dict() if len(df) > 0 else {{}},
    "null_counts": df.isnull().sum().to_dict(),
    "sample": df.head(5).to_dict('records')
}}

print(json.dumps(analysis))
'''
        
        result = await self.sandbox.execute_code(code)
        
        if not result.get("success"):
            return {"success": False, "error": result.get("error")}
        
        try:
            analysis = json.loads(result.get("stdout", "{}"))
            return {"success": True, "analysis": analysis}
        except:
            return {"success": False, "error": "Failed to parse analysis results"}


class AnalystLayer:
    """
    The V3.1 Analyst Layer
    Combines E2B sandbox with specialized generators for professional output.
    """
    
    def __init__(self):
        self.sandbox = E2BSandbox()
        self.excel_gen = ExcelGenerator(self.sandbox)
        self.analyzer = DataAnalyzer(self.sandbox)
    
    async def execute(self, query: str, context: Dict = None) -> Dict:
        """
        Execute an analyst task based on the query and context.
        """
        
        context = context or {}
        task_type = context.get("task_type", "analysis")
        data = context.get("data", {})
        
        files = []
        analysis = ""
        
        if task_type == "excel" or "excel" in query.lower() or "spreadsheet" in query.lower():
            # Generate Excel file
            title = context.get("title", "Fashion Report")
            result = await self.excel_gen.generate(data, title)
            
            if result.get("success"):
                files.append({
                    "filename": result["filename"],
                    "data": result["file_data"],
                    "content_type": result["content_type"]
                })
                analysis = f"Generated Excel report: {result['filename']}"
            else:
                analysis = f"Failed to generate Excel: {result.get('error')}"
        
        elif task_type == "analyze" or "analyze" in query.lower():
            # Perform data analysis
            result = await self.analyzer.analyze(data)
            
            if result.get("success"):
                analysis = json.dumps(result["analysis"], indent=2)
            else:
                analysis = f"Analysis failed: {result.get('error')}"
        
        else:
            # Custom code execution
            code = context.get("code", "")
            if code:
                result = await self.sandbox.execute_code(code)
                analysis = result.get("stdout", "") or result.get("error", "No output")
        
        return {
            "success": True,
            "analysis": analysis,
            "files": files
        }
    
    async def generate_excel_from_search(self, search_results: List[Dict], title: str) -> Dict:
        """Generate an Excel file from search results."""
        
        data = {
            "columns": ["Title", "URL", "Source", "Snippet"],
            "rows": [
                [r.get("title", ""), r.get("url", ""), r.get("source", ""), r.get("snippet", "")[:200]]
                for r in search_results
            ]
        }
        
        return await self.excel_gen.generate(data, title)


# Global instance
analyst_layer = AnalystLayer()
